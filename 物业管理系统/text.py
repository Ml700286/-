import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def create_test_cases():
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()

    # ====== 功能测试表 ======
    func_sheet = wb.active
    func_sheet.title = "功能测试"

    # 设置表头
    headers = ["用例编号", "测试模块", "操作步骤", "输入数据", "预期结果", "实际结果", "是否通过"]
    func_sheet.append(headers)

    # 添加管理员功能测试用例
    admin_cases = [
        {
            "模块": "管理员登录",
            "步骤": "1. 输入用户名和密码\n2. 点击登录",
            "输入": "用户名: admin\n密码: admin",
            "预期": "登录成功，进入管理员主界面"
        },
        {
            "模块": "添加业主",
            "步骤": "1. 进入业主管理\n2. 填写业主信息\n3. 点击保存",
            "输入": "用户名: test_user\n密码: 123456\n车牌: 津A88888",
            "预期": "提示'添加成功'，custom_account表新增记录"
        },
        {
            "模块": "管理房产信息",
            "步骤": "1. 进入房产管理\n2. 添加新房源信息\n3. 点击保存",
            "输入": "门牌号: A-1001\n楼号: 5\n类型: 高层\n业主ID: 1",
            "预期": "house表新增记录，外键关联正确"
        },
        {
            "模块": "处理报修单",
            "步骤": "1. 进入报修管理\n2. 选择未处理报修\n3. 更新状态和实际花费\n4. 保存",
            "输入": "报修ID: 1\n状态: 已完成\n实际花费: 150.00",
            "预期": "maintain表状态更新，scost字段更新"
        },
        {
            "模块": "发布公告",
            "步骤": "1. 进入公告管理\n2. 填写公告内容\n3. 点击发布",
            "输入": "标题: 停水通知\n内容: 明天上午停水检修\n发布者: 物业部",
            "预期": "notice表新增记录，业主端可见新公告"
        }
    ]

    for i, case in enumerate(admin_cases, 1):
        func_sheet.append([
            f"TC-ADM-{i:03d}",
            case["模块"],
            case["步骤"],
            case["输入"],
            case["预期"],
            "",  # 实际结果留空
            ""  # 是否通过留空
        ])

    # 添加业主功能测试用例
    owner_cases = [
        {
            "模块": "业主登录",
            "步骤": "1. 输入用户名和密码\n2. 点击登录",
            "输入": "用户名: ysj\n密码: 1234",
            "预期": "登录成功，进入业主主界面"
        },
        {
            "模块": "查看个人信息",
            "步骤": "1. 进入个人中心",
            "输入": "",
            "预期": "显示业主用户名、车牌信息"
        },
        {
            "模块": "查看房产信息",
            "步骤": "1. 进入我的房产",
            "输入": "",
            "预期": "显示关联的房产信息(门牌号、楼号等)"
        },
        {
            "模块": "查看报修记录",
            "步骤": "1. 进入报修记录",
            "输入": "",
            "预期": "显示当前业主的报修历史记录"
        },
        {
            "模块": "查看社区公告",
            "步骤": "1. 进入公告栏",
            "输入": "",
            "预期": "显示所有已发布的社区公告"
        }
    ]

    for i, case in enumerate(owner_cases, 1):
        func_sheet.append([
            f"TC-OWN-{i:03d}",
            case["模块"],
            case["步骤"],
            case["输入"],
            case["预期"],
            "",  # 实际结果留空
            ""  # 是否通过留空
        ])

    # ====== 数据库验证表 ======
    db_sheet = wb.create_sheet("数据库验证")
    db_headers = ["用例编号", "测试动作", "SQL验证语句", "预期数据状态", "实际数据状态", "是否一致"]
    db_sheet.append(db_headers)

    db_cases = [
        {
            "动作": "添加业主后",
            "sql": "SELECT * FROM custom_account WHERE username='test_user'",
            "预期": "返回1条记录，车牌为'津A88888'"
        },
        {
            "动作": "添加房产后",
            "sql": "SELECT * FROM house WHERE num='A-1001' AND ownerid=1",
            "预期": "返回1条记录，dep='5', type='高层'"
        },
        {
            "动作": "更新报修单后",
            "sql": "SELECT status, scost FROM maintain WHERE id=1",
            "预期": "status='已完成', scost=150.00"
        },
        {
            "动作": "发布公告后",
            "sql": "SELECT * FROM notice WHERE title='停水通知'",
            "预期": "返回1条记录，content包含'停水检修'"
        },
        {
            "动作": "业主查看房产",
            "sql": "SELECT COUNT(*) FROM house WHERE ownerid=(SELECT ownerid FROM custom_account WHERE username='ysj')",
            "预期": "返回大于0条记录"
        }
    ]

    for i, case in enumerate(db_cases, 1):
        db_sheet.append([
            f"TC-DB-{i:03d}",
            case["动作"],
            case["sql"],
            case["预期"],
            "",  # 实际数据状态留空
            ""  # 是否一致留空
        ])

    # ====== 异常流测试表 ======
    exc_sheet = wb.create_sheet("异常流测试")
    exc_headers = ["用例编号", "异常场景", "非法输入", "预期系统提示", "实际提示", "是否通过"]
    exc_sheet.append(exc_headers)

    exc_cases = [
        {
            "场景": "使用错误密码登录(管理员)",
            "输入": "用户名: admin\n密码: wrong",
            "预期": "提示'用户名或密码错误'"
        },
        {
            "场景": "添加重复门牌号",
            "输入": "门牌号: A-1001\n(已存在相同门牌号)",
            "预期": "提示'门牌号已存在'"
        },
        {
            "场景": "业主查看他人房产",
            "输入": "尝试访问非关联房产ID",
            "预期": "提示'无权查看该房产信息'或返回空数据"
        },
        {
            "场景": "输入负数的维修费用",
            "输入": "实际花费: -100",
            "预期": "提示'费用不能为负数'"
        },
        {
            "场景": "删除不存在的业主",
            "输入": "业主ID: 9999",
            "预期": "提示'业主不存在'"
        }
    ]

    for i, case in enumerate(exc_cases, 1):
        exc_sheet.append([
            f"TC-EXC-{i:03d}",
            case["场景"],
            case["输入"],
            case["预期"],
            "",  # 实际提示留空
            ""  # 是否通过留空
        ])

    # ====== 样式美化 ======
    # 设置列宽
    for sheet in wb:
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 35
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 35
        sheet.column_dimensions['F'].width = 35
        sheet.column_dimensions['G'].width = 10

        # 设置标题行样式
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        border = Border(bottom=Side(style='medium'))

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # 设置自动换行
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    # 保存文件
    wb.save("物业管理系统_测试用例.xlsx")
    print("测试用例表已生成: 物业管理系统_测试用例.xlsx")


if __name__ == "__main__":
    create_test_cases()